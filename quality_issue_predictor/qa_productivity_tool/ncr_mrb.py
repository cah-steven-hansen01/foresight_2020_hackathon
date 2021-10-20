# Standard Libraries
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side
import time
import datetime as dt
import matplotlib.pyplot as plt
import os

# Custom Libraries
import CleaningTools as ct
import Reference
from quality_report import Quality_Report
# Plot styles
plt.style.use('ggplot')
['Solarize_Light2', '_classic_test_patch', 'bmh', 'classic', 'dark_background', 'fast', 'fivethirtyeight', 'ggplot', 'grayscale', 'seaborn', 'seaborn-bright', 
'seaborn-colorblind', 'seaborn-dark', 'seaborn-dark-palette', 'seaborn-darkgrid', 'seaborn-deep', 'seaborn-muted', 'seaborn-notebook', 'seaborn-paper', 'seaborn-pastel', 'seaborn-poster', 'seaborn-talk', 'seaborn-ticks', 'seaborn-white', 'seaborn-whitegrid', 'tableau-colorblind10']

## tests

def test_run_report():
    report = NCR_MRB()
    clean_data_df,task_count_df =  report.run_report()
    clean_data_df.to_clipboard()
    #task_count_df.to_clipboard()
def test_clipboard_data():
    report = NCR_MRB()
    #report.clipboard_data('task_count')
    report.clipboard_data('clean')
def test_default_report():
    filename = 'NCR_MRB_test_default_report' + str(time.time()).split('.')[0]
    report = NCR_MRB()
    report.default_report(filename=filename,openfile='y')
def test_metrics():
    report = NCR_MRB()
    report.mostrecentreport()
    metrics =report.metrics()
    print(metrics)
def test_visual_age_distribution():
    report = NCR_MRB()
    report.run_report()
    report.visual_age_distribution()
def test_visual_NC_age_bar():
    report = NCR_MRB()
    report.mostrecentreport()
    report.run_report()
    report.visual_NC_age_bar()
def test_visual_open_nc_table():
    report = NCR_MRB()
    report.visual_open_nc_table()
def test_visual_tierII_tracker():
    report = NCR_MRB()
    report.visual_tierII_tracker()
def test_run_report():
    report = NCR_MRB()
    report.mostrecentreport()
    report.run_report()[0].to_clipboard()
protocol = test_run_report


class NCR_MRB(Quality_Report):
    ''' Cleans the report NCR_MRB from SmartSolve and outputs clean data'''
    def __init__(self):
        super().__init__(report_name='NC Report - MRB')
        #self.r_filename, self.path, self.date = ct.mostrecentreport("NC Report - MRB")
        #self.status_of_report = 0 # becomes 1 when run_report performed
        self.task_timing_dict = {
            'Verify Non Conformance Initiation': 1,
            'Containment Action':2,
            'Investigate Non Conformance': 7,
            'Approve Investigation Level 1': 8,
            'Sign-off Action Plans': 25,
            'Plan Disposition': 27,
            'Approve Disposition Plan': 27,
            'Execute Disposition': 27,
            'Adhoc':28,
            'Verify Non Conformance Closure': 29
        }
        self.ref = Reference.Reference()
    # def meta_data(self):
    #     return self.r_filename, self.path, self.date
    def metrics(self):
        if self.status_of_report == 0:
            self.run_report()
        self.num_of_open = len(self.clean_data_df['NC Number'])
        self.num_over_60 = sum(self.clean_data_df['Over 60 days'][self.clean_data_df['Over 60 days']==1])
        self.percent_open_under_60d = round((self.num_of_open-self.num_over_60)/self.num_of_open,3)
        
        return self.num_of_open, self.percent_open_under_60d          
    def run_report(self):
        '''outputs a clean_data_df and a count of the tasks'''
        print("\nRunning NCR MRB Report... ")
        #self.mostrecentreport()
        self.r_data = pd.read_excel(open(self.path,'rb'))
        # Overall Formatting
        self.data = self.r_data.iloc[4:]
        self.data = self.data.drop(4)
        self.data = self.data.rename(columns = self.data.iloc[0,:])
        self.data = self.data.drop(5)
        # Date formating
        self.data["Created Date"] = pd.to_datetime(self.data["Created Date"])
        self.data["Created Date"] = self.data["Created Date"].dt.date
        # Product code formatting.
        self.data["Product"] = list(map(self.tmpecc_remover,self.data["Product"]))
        # get rid of the 'Total Number of records' useless line item
        self.data = self.data[self.data["NC Number"] != self.data.iloc[-1,0]]
        self.data = self.data.reset_index(drop=True)
        self.copy_data = ct.phase_compiler(self.data) # copy for tier II tracker
        self.clean_data_df = self.data
        self.clean_data_df = self.task_timing() 

        self.clean_data_df = ct.phase_compiler(self.clean_data_df)
        self.clean_data_df = ct.index_compiler("NC Number",self.clean_data_df,retain_duplicates=['Total Quantity Affected'])
        # create "Total Qty" column
        self.clean_data_df['Total Qty']=0
        for i, qty in enumerate(self.clean_data_df['Total Quantity Affected']):
            try:
                self.clean_data_df.loc[i,'Total Qty'] = sum([int(q) for q in qty.strip('[ ]').split(', ')])
            except AttributeError:
                self.clean_data_df.loc[i,'Total Qty'] = qty
            except ValueError:
                pass
        # Add cost
        self.clean_data_df = self.product_cost()
        # create 'Over 60 day' column
        self.clean_data_df['Over 60 days']=0
        for i, age in enumerate(self.clean_data_df['Age of NC']):
            if age >= 60:
                self.clean_data_df.loc[i,'Over 60 days'] = 1
        self.task_count = ct.task_counterII(self.clean_data_df)
        self.clean_data_df = self.clean_data_df.dropna(subset=['NC Number'])

        self.clean_data_df = self.clean_data_df.reset_index(drop=True)
        # product family / value stream
        
        self.clean_data_df = self.ref.evaluate_pc(self.clean_data_df)
        # remove nan column
        self.clean_data_df = self.clean_data_df.loc[:,self.clean_data_df.columns.notnull()]
        
        self.status_of_report = 1
        three_days_ago = dt.datetime.now().date() - dt.timedelta(days = 3)
        cols_for_recent_ncs = ["NC Number",'Created Date','Initial Failure Mode','Value Stream']
        self.nc_in_last_72hrs = self.clean_data_df[cols_for_recent_ncs][self.clean_data_df['Created Date'] > three_days_ago]
        if len(self.nc_in_last_72hrs) == 0:
            print('No new NCs have been opened in the last 72 hrs.')
        else:
            print("\nNCs opened in the last 72hrs:\n ",self.nc_in_last_72hrs)
        self.metrics()
        print('\nNumber of open NC = ', self.num_of_open)
        print('Percent open under 60 days = ',self.percent_open_under_60d)
        return self.clean_data_df, self.task_count
    def ncs_last_72hrs(self):
        return self.nc_in_last_72hrs
    def task_timing(self):
        # Creates column of overdue tasks
        for i, task in enumerate(self.data['Task Name']):
            # print(self.task_timing[task])
            # print(self.data.loc[i,'Age of NC'])
            if task == 'Plan Disposition':
                pass
            elif int(self.data.loc[i,'Age of NC']) > self.task_timing_dict[task]:
                days_over_due = str(self.data.loc[i,'Age of NC'] - self.task_timing_dict[task])
                self.data.loc[i,"Overdue Task"] = str('"'+self.data.loc[i,'Task Name']+'"' +' is '+ 
                days_over_due + ' days over due and assigned to ' + self.data.loc[i,'Task Owner']+' ')
        return self.data
    def product_cost(self):
        for i, pc in enumerate(self.clean_data_df['Product']):
            #print(ref.evaluate_cost(pc) )
            self.clean_data_df.loc[i,'Cost'] = self.ref.evaluate_cost(pc) * self.clean_data_df.loc[i,'Total Qty']
        return self.clean_data_df
    def visual_age_distribution(self,save_to_path = './Figures',show = False,for_embed = False):
        if self.status_of_report == 0:
            self.run_report()
        ### scaling the Over 60 Day subplot ###
        self.metrics()
        over60_axis_ratio = 1-self.percent_open_under_60d
        #over60_axis_ratio = 0
        if over60_axis_ratio == 0: # run the plot that doesn't have subplots
            self._visual_dist_no_over_60(save_to_path = save_to_path)
            return None
        elif over60_axis_ratio<0.25: # too small looks ridiculous
            f, (axes) = plt.subplots(ncols=2,nrows=1,figsize = (10, 5),sharey = True,gridspec_kw={'width_ratios': [1, .25]})
        else: # use ratio
            f, (axes) = plt.subplots(ncols=2,nrows=1,figsize = (10, 5),sharey = True,gridspec_kw={'width_ratios': [1, over60_axis_ratio]})
        ##############################################
        bins = [0,7,14,21,25,30,35,40,45,50,55,60]
        alertlevel = ['g','g','g','yellow','yellow','r','r','r','r','r','r']
        N,bins, patches = axes[0].hist(x = self.clean_data_df['Age of NC'],bins = bins,rwidth = .9)
        for alertcolor,thispatch in zip(alertlevel,patches):
            thispatch.set_facecolor(alertcolor)
        fontsize = 9
        axes[0].set_title('Age Distribution of Open NCs\n Under 60 Days',fontsize=fontsize)
        axes[0].set_xticks(bins)
        over60daysNCs= self.clean_data_df[self.clean_data_df['Age of NC'] >60]
        axes[1].hist(x = over60daysNCs['Age of NC'],facecolor = 'r',rwidth = .9)
        axes[1].set_title('Age Distribution of Open NCs\n Over 60 days',fontsize=fontsize)
        
        
        #f.text(0.5, 0.04, 'Age', fontsize = fontsize,ha='center', va='center')
        #f.text(.08,0.5,'Number of NCs',fontsize = fontsize,ha = 'center',va='center',rotation = 'vertical')
        
        #f.suptitle('Age Distribution of NCs',fontsize = 16)
        f.tight_layout() # messes up the axes labelling
        if show == True:
            plt.show()
        f.savefig(save_to_path+r'\\'+'Distribution of NC Age')
        plt.close(f)
    def visual_NC_age_bar(self,save_to_path='./Figures',show = False):
        if self.status_of_report == 0:
            self.run_report()
        num_of_nc = len(self.clean_data_df['NC Number'])
        thirty_day_line = num_of_nc * [30]
        sixty_day_line = num_of_nc * [60]
        
        plt.bar(x= self.clean_data_df['NC Number'], height = self.clean_data_df['Age of NC'],color = 'b')
        plt.plot(self.clean_data_df['NC Number'],thirty_day_line,'k',label = '30 day line')
        plt.plot(self.clean_data_df['NC Number'],sixty_day_line,'r',label = '60 day line')
        plt.tick_params(labelrotation = 90)
        plt.title("Age of Open NCs")
        plt.legend()
        plt.tight_layout()
        plt.savefig(save_to_path+r'\\'+'Age of Open NCs')
        if show == True:
            plt.show()
        plt.close()
        
        
        
        #plt.close()
    def _visual_dist_no_over_60(self, save_to_path='./Figures'):
        '''same plot as visual_age_distribution but will be used when no NCs are 
        over 60 days'''
        f, (axes) = plt.subplots(ncols=1,nrows=1,figsize = (10, 5))
        bins = [0,7,14,21,25,30,35,40,45,50,55,60]
        alertlevel = ['g','g','g','yellow','yellow','r','r','r','r','r','r']
        N,bins, patches = axes.hist(x = self.clean_data_df['Age of NC'],bins = bins,rwidth = .9)
        for alertcolor,thispatch in zip(alertlevel,patches):
            thispatch.set_facecolor(alertcolor)
        axes.set_title('Age Distribution of Open NCs')
        axes.set_xticks(bins)
        f.tight_layout()
        #plt.show()
        f.savefig(save_to_path+r'\\'+'Distribution of NC Age')
        plt.close(f)
        
        print("no NCs over 60 days")
        pass
    def visual_open_nc_table(self,save_to_path = './Figures'):
        if self.status_of_report == 0:
            self.run_report()
        import openpyxl
        
        data = self.clean_data_df[['NC Number','Age of NC','Initial Failure Mode',
        'Task Name','Task Owner','Total Qty','Overdue Task']] 
        filename = "Open NC Phase Tracker "+dt.datetime.now().strftime('%d-%b-%Y')
        report = openpyxl.Workbook()
        report.create_sheet(index=0,title = 'Open NCs Tracker')
        tracker_sheet = report.active
        ct.copy_paste_sheet(data,tracker_sheet)
        col_widths = {'A': 15, 'B': 10, 'C': 41, 'D': 54, 'E': 50, 'F': 9, 'G': 85}
        for col,w in col_widths.items():
            tracker_sheet.column_dimensions[col].width = w
        report.save(save_to_path+r'\\'+filename+'.xlsx')   
    def visual_tierII_tracker(self,save_to_path='./Figures',openfile = 'n'):
        '''Outputs the tierII_tracker'''
        if self.status_of_report == 0:
            self.run_report()
        data = self.copy_data
        data = data[['NC Number','Age of NC','Initial Failure Mode','Description',
        'Verify Non Conformance Initiation','Containment Action','Investigate Non Conformance','Approve Investigation Level 1','Sign-off Action Plans',
        'Plan Disposition','Approve Disposition Plan','Execute Disposition','Adhoc','Verify Non Conformance Closure']]
        data = ct.index_compiler('NC Number',data)
        for phase,due_day in self.task_timing_dict.items():
            if not phase == 'Plan Disposition':
                for i, person in enumerate(data[phase]):
                    if not str(person) == 'nan':
                        if data.loc[i,'Age of NC'] > due_day:
                            days_over_do = data.loc[i,'Age of NC'] - due_day
                            data.loc[i,phase] = person + ' Task Past Due!'# ('+str(days_over_do)+' days over)'
                        else:
                            days_till_due = due_day - data.loc[i,'Age of NC']
                            data.loc[i,phase] = person + ' Task Due in '+str(days_till_due)+' days.'
        report = openpyxl.Workbook()
        report.create_sheet(index=0,title = 'TierII Tracker')
        tracker_sheet = report.active
        ct.copy_paste_sheet(data,tracker_sheet)
        col_widths = {'A': 15, 'B': 9, 'C': 23, 'D': 78, 
        'E': 14, 'F': 14, 'G': 14,'H':14,'I':14,'J':14,
        'K':14,'L':14,'M':14,'N':14}
        for col,w in col_widths.items():
            tracker_sheet.column_dimensions[col].width = w
        for row in tracker_sheet.iter_rows():  
            for cell in row:      
                cell.alignment = Alignment(wrap_text=True,vertical='top') 
        
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        fill_red = PatternFill("solid", fgColor="00FF0000")
        white_font = Font(color='00FFFFFF')
        # standard 14 columns
        for i in range(len(data['NC Number'])+1):
            i +=1
            for j in range(len(col_widths.keys())):
                j+=1
                tracker_sheet.cell(row = i,column = j).border = thin_border
                cell_value = tracker_sheet.cell(row = i,column = j).value
                if 'Past Due!' in str(cell_value):
                    tracker_sheet.cell(row = i,column = j).fill = fill_red
                    tracker_sheet.cell(row = i,column = j).font = white_font
        last_row = len(data['NC Number'])+2
        tracker_sheet.cell(row = last_row,column = 1 ).value = "Original Data = "
        tracker_sheet.cell(row = last_row,column = 2 ).value = self.r_filename
        tracker_sheet.oddHeader.center.text = 'Tier II Tracker ' + dt.datetime.now().strftime('%d-%b-%Y')
        tracker_sheet.oddHeader.center.size = 18
        filename = "NC Task Tracker "+dt.datetime.now().strftime('%d-%b-%Y_t%H%M') 
        ## saving
        file_path = save_to_path+r'\\'+filename+'.xlsx'
        try:
            report.save(file_path)
        except PermissionError:
            import random
            rand_num = random.randint(1,1000)
            file_path = save_to_path+r'\\'+filename+str(rand_num)+'.xlsx'
            report.save(file_path)
            print('Permission Error - Filname changed to '+filename+'_'+str(rand_num))
        if openfile == 'y':
            os.startfile(file_path)

    def default_report(self,filename,openfile = 'n',to_path=r"C:\\Users\\steven.hansen01\\data_automation\\venv\\Test_Reports\\"):
        '''This will output an excel report'''
        if self.status_of_report == 0:
            self.run_report() 
        print('Building Report...')
        import os
        self.filename = filename
        self.report = openpyxl.Workbook()
        # meta_data
        self.report.create_sheet(index = 0, title = "meta_data ")
        self.meta_sheet = self.report.active
        self.meta_sheet["A1"] = "Original Data Filename"
        self.meta_sheet["A2"] = self.r_filename
        self.meta_sheet["B1"] = "Data generated: "
        self.meta_sheet["B2"] = self.date
        self.meta_sheet.column_dimensions["A"].width = 30
        # raw_data
        self.report.create_sheet(index=0,title = "raw_data")
        self.raw_data_ws = self.report.active
        ct.copy_paste_sheet(self.r_data,self.raw_data_ws)
        # clean_data
        self.report.create_sheet(index = 0,title="clean_data")
        self.clean_data_ws = self.report.active
        ct.copy_paste_sheet(self.clean_data_df,self.clean_data_ws)
        cols_width = {'A':16,'B':12.30,'C':14,'D':32,'E':32,'F':32,'G':5,'H':45,'I':24,'J':2,"K":24,"L":22,"M":25,"N":25,"O":9}
        ct.excel_col_width(self.clean_data_ws,cols_width)
        ct.excel_text_wrap(self.clean_data_ws,[1,15])
        # visual_data
        self.report.create_sheet(title = "visual_data",index=0)
        self.visual_data = self.report.active
        ct.excel_bar_chart(ref_sheet=self.clean_data_ws,dis_sheet=self.visual_data,
        title = "NC Age",y_axis_title="Age",x_axis_title='NC',y_col_num=15,x_col_num=1,cell = 'A1')
        # save and open (optional)
        self.report.save(to_path+self.filename+'.xlsx')
        if openfile == 'y':
            os.startfile(to_path+self.filename+'.xlsx')






if __name__ == '__main__':
    protocol()

    


