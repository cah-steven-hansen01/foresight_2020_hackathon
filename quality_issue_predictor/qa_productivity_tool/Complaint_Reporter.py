import os
import pandas as pd
rootpath = r"C:\Users\steven.hansen01\Documents\Python_Development"
os.chdir(rootpath)
pd.options.mode.chained_assignment = None
import openpyxl
import numpy as np
from openpyxl.chart import BarChart, Series, Reference

print("Complaint Reporter Version 1.0\nThis is not a validated program and is intended for information/demostration purposes only\n\n")

def mostrecentReport(report):
    entries = {}
    path = r"C:\Users\steven.hansen01\Downloads"
    with os.scandir(path) as it:
        for entry in it:
            if entry.name.startswith(report):
            #print(entry.name)
                entries.update({entry.name:[entry.stat().st_ctime,entry.path]})
    for key in entries:
        if entries[key] == max(entries.values()):
            timepulled = pd.Timestamp(entries[key][0], unit = 's')
            day,month,year = timepulled.day,timepulled.month,timepulled.year
            datepulled = str(month)+"/"+str(day)+"/"+str(year)
            return key,entries[key][1], datepulled

runtheprogram = input("Run program? y/n ")
if runtheprogram == 'n':
    exit()
    
     
filename, path,complaintpulled = mostrecentReport("Case_CAH_By_Customer_Name")
data=pd.read_excel(open(path,'rb'))
raw_data = data
print("Most recent Complaint report found  = ",filename)
print("Report pulled: ",complaintpulled)


filename, path, regpulled = mostrecentReport("Regulatory Reporting Global Report")
reg_data=pd.read_excel(open(path,'rb'))
raw_reg_data = reg_data
print("Most recent Regulatory report found  = ",filename)
print("Report pulled: ",regpulled)

filename, path,cmplnt_task_pulled = mostrecentReport("Complaint Average Time to Close")
cmplnt_task_data = pd.read_excel(open(path,'rb'))
raw_cmplnt_task_data = cmplnt_task_data
print("Most recent Complaint Task report found  = ",filename)
print("Report pulled: ",cmplnt_task_pulled)

# Cross reference Clean up
cross_ref = pd.read_csv("./Reference/Cross_Reference.csv")
fg = cross_ref["Family"].drop([0]) #FGs are in the same column as the header for the section
family = cross_ref[cross_ref.columns[1]].drop([0])
lids = cross_ref[cross_ref.columns[2]].drop([0])
ff = cross_ref[cross_ref.columns[3]].drop([0])
pc_pf_ref = {"Finished Good":fg,"Family":family,"Lid":lids,"Focus Factory":ff}
pc_pf_ref = pd.DataFrame(pc_pf_ref)
pc_pf_ref = pc_pf_ref.dropna(subset = ["Finished Good"])
pc_pf_ref = pc_pf_ref.reset_index()
risk = cross_ref.iloc[:,11:14]
risk = risk.rename(columns = {"Unnamed: 11":"Failure Mode","Unnamed: 12":"Failure Family","Unnamed: 13":"Risk"})
risk = risk.drop([0]) 

# Clean up of Complaint Task report
for i,col in enumerate(cmplnt_task_data.columns):
    cmplnt_task_data = cmplnt_task_data.rename(columns = {col:cmplnt_task_data.iloc[7,i]})
cmplnt_task_data = cmplnt_task_data.drop([0,7])
cmplnt_task_data = cmplnt_task_data.loc[:,cmplnt_task_data.columns.notnull()]
cmplnt_task_data = cmplnt_task_data.dropna(subset = ["Record Number"])
cmplnt_task_data = cmplnt_task_data.reset_index()
cmplnt_task_data = cmplnt_task_data.drop(["index"],axis=1)
phases = {}
for i, record in enumerate(cmplnt_task_data['Record Number']):
    if record.split('-')[0] != "CASE":
        phases.update({record:i})

cmplnt_task_data['Phase'] = 0
for k in phases:
    cmplnt_task_data["Phase"].iloc[phases[k]:] = k
invest_task = cmplnt_task_data[cmplnt_task_data["Phase"]=="Investigate Complaint"]
approve_invest_task = cmplnt_task_data[cmplnt_task_data["Phase"]=="Approve Complaint Investigation"]

# Format Datetime
raw_data['Event Date']=pd.to_datetime(raw_data['Event Date'])
raw_data['Complaint Create Date']=pd.to_datetime(raw_data['Complaint Create Date'])
raw_data['Manufactured Date']=pd.to_datetime(raw_data['Manufactured Date'])
raw_data['Case Closed Date']=pd.to_datetime(raw_data['Case Closed Date'])
raw_data['Complaint Closure Date']=pd.to_datetime(raw_data['Complaint Closure Date'])
raw_data['Sample Received Date']=pd.to_datetime(raw_data['Sample Received Date'])

data['Event Date']=pd.to_datetime(data['Event Date'])
data['Complaint Create Date']=pd.to_datetime(data['Complaint Create Date'])
data['Case Closed Date']=pd.to_datetime(data['Case Closed Date'])

# function for reviewing column names during inital data cleaning 
def see_values(column_name):
    values=data[column_name][data[column_name].isna()==False]
    num_of_values = len(values)
    print(f'number of values is {num_of_values}')
    print(values)

# Drop columns with 100% NAN and other not useful columns
allna = ["Product Group","Reported Component", "Reported Component Name","Reported Vendor Name"]
useless_columns = ["Note Type", "Sales Rep","Sales Support","Customer Service"]
columns_low_use = ["Manufactured Date","Investigated Component","Investigated Component Name","Investigated Component Lot Number",
                  "Investigated Mfg. SKU Number","Investigated Vendor Name",]
columns_to_drop = allna + useless_columns+columns_low_use
data = data.drop(columns_to_drop,axis=1)
# Clean up the ~TPMECC
i=0
while i<len(data):
    data.Product[i] = data.Product[i].split('~')[0]
    i+=1
# Regulatory clean up
reg_data = reg_data.iloc[7:,0]
reg_data = reg_data.reset_index()
reg_data = reg_data.drop('index',axis=1)
reg_data = reg_data.rename(columns = {reg_data.columns[0]:"Record Number"})
reg_data = reg_data.dropna()

## translates from "Record Number" to "Complaint Number" in regulatory report
reg_data["Complaint Number"] = 0
for i, record in enumerate(reg_data["Record Number"]):
    reg_data["Complaint Number"][i] = record[0:20]   
## formats as sets and creats set of cl reportable complaints
reg_set = set(reg_data["Complaint Number"])
cl_complaint_set = set(data["Complaint Number"])
cl_reportable = reg_set&cl_complaint_set
## Looks up and assigns regulatory status 
data["Reportable"] = 0
for i, complaint in enumerate(data["Complaint Number"]):    
    if complaint in cl_reportable:
        data["Reportable"][i] = "Yes"
    else:
        data["Reportable"][i] = "No"
for i, date in enumerate(data['Complaint Create Date']):
    if date> pd.to_datetime(regpulled):
        data["Reportable"][i] = "Complaint created after most recent regulatory report data"

        # Look up functions
def pf_lookup(pc):
    return str(pc_pf_ref[pc_pf_ref["Finished Good"]==pc]["Family"]).split()[1]
def risk_lookup(fm):
    if str(fm)=='nan':
        return "No Investigated Failure Mode"
    else:
        return str(risk[risk["Failure Mode"]==fm]["Risk"]).split()[1]
def factory_lookup(pc):
    return str(pc_pf_ref[pc_pf_ref["Finished Good"]==pc]["Focus Factory"]).split()[1]

i=0
Product_Family = []
Factory = []
Risk = []

while i<len(data["Product"]):
    if data["Product"][i] == "065":
        data["Product"][i]="65"
    Product_Family.append(pf_lookup(data["Product"][i]))
    Factory.append(factory_lookup(data["Product"][i]))
    if risk_lookup(data["Investigated Failure Mode"][i]) == "Name:":
        Risk.append("Not Yet Evaluated")
    else:
        Risk.append(risk_lookup(data["Investigated Failure Mode"][i]))
    i+=1
data["Risk"] = pd.Series(Risk,index = data.index)
data["Value Stream"] = pd.Series(Factory, index=data.index)
data["Product Family"] = pd.Series(Product_Family, index=data.index)
fm_nofound = len(data['Risk']=="Not Yet Evaluated")
num_pc_notfound = len(data['Product'][data["Product Family"]=="Name:"])
print(F'{num_pc_notfound} product codes were not found')
if num_pc_notfound>0:
    print(data['Product'][data["Product Family"]=="Name:"])

# Calculates overal complaint age
data["Age"] = 0
for i, c in enumerate(data['Complaint Create Date']):
    if str(data["Case Closed Date"][i]) == "NaT":
        data["Age"][i] = str(pd.to_datetime('today') - data['Complaint Create Date'][i]).split()[0]
    else:
        data["Age"][i] = str(data["Case Closed Date"][i]-data['Complaint Create Date'][i]).split()[0]

# standardizes columns
std_columns = ['Complaint Create Date','Complaint Number','Product','Lot # / Work Order','Product Description Summary',
                 'Reported Failure Mode',"Investigated Failure Mode",'Investigation Summary','Reportable','Case Closed Date',
               'Risk',"Value Stream","Product Family","Age"]

Open_complaint = data[data["Complaint Closure Date"].isna()]
Open_complaint = Open_complaint[std_columns]
Open_complaint = Open_complaint.reset_index()
Open_complaint = Open_complaint.drop(['index'],axis=1)

# Checks submission status for open investigations
complaints = []
status = []
invest_submit_date = []
for i, complaint in enumerate(Open_complaint["Complaint Number"]):
    try:
        if max(approve_invest_task["Sign-off Date"][approve_invest_task["Record Number"]==complaint])>max(invest_task["Sign-off Date"][invest_task["Record Number"]==complaint]):
            # this looks at if the apprval task or investigate task is more recent
            complaints.append(complaint)
            status.append("Rejected")
            invest_submit_date.append(max(invest_task["Sign-off Date"][invest_task["Record Number"]==complaint]))
        else:
            complaints.append(complaint)
            status.append("Resubmitted")
            invest_submit_date.append(max(invest_task["Sign-off Date"][invest_task["Record Number"]==complaint]))

    except:
        # potential bug - if investigation summary is filled out but no investigation is submitted.  
        # theoritically will error out if that is the case because no invest_task["Sign-off Date"] would exist
        # Bug resolved!
        if str(Open_complaint["Investigation Summary"][i])=="nan":
            complaints.append(complaint)
            status.append("No Submission")
            invest_submit_date.append("No Submission")
            
        else:
            try:
                complaints.append(complaint)
                
                invest_submit_date.append(max(invest_task["Sign-off Date"][invest_task["Record Number"]==complaint]))
                status.append("Waiting for Inv Approval")
            except:
                invest_submit_date.append("sign-off data not confirmed")
                status.append("Inv summary inputed, but inv may not be submitted or task report is not updated")
            
results = {"Complaint Number":complaints,"Submission Status":status,"Date of Investigation":invest_submit_date}
results_df = pd.DataFrame(results)
Open_complaint = pd.merge(Open_complaint, results_df,on='Complaint Number')


Trend_data = data[["Product Family",'Complaint Create Date',"Complaint Number","Product",'Lot # / Work Order',
                   "Investigated Failure Mode","Complaint Quantity"]]
## Calculates TAT for Cleaned_data sheet

path = "./Reference/site_CMPLNT_TAT.xlsx"
site_CMPLNT_TAT = pd.read_excel(open(path,'rb'))

Cleaned_data = data[std_columns]
site_CMPLNT_TAT_df = site_CMPLNT_TAT[["Complaint Number","Site TAT"]]
Cleaned_data = pd.merge(Cleaned_data,site_CMPLNT_TAT_df,on="Complaint Number",how = "left")
Cleaned_data['TAT Source'] = "Manual Lookup"
investigate = cmplnt_task_data[cmplnt_task_data["Phase"]=="Investigate Complaint"]
for i,complaint in enumerate(Cleaned_data["Complaint Number"]):
    if (Cleaned_data["Site TAT"][i]==0) or (str(Cleaned_data["Site TAT"][i])=='nan'):
        if str(Cleaned_data["Investigated Failure Mode"][i]) == 'nan':
            Cleaned_data["Site TAT"][i] = ''
            Cleaned_data['TAT Source'][i] = 'No Investigation submitted'
        else:
            Cleaned_data["Site TAT"][i] =investigate['Complaint Task Age'][investigate["Record Number"] == complaint].sum()
            Cleaned_data['TAT Source'][i] = "Complaint Task Report"
            

open_cl_inv = Open_complaint[["Complaint Number","Date of Investigation","Submission Status"]]
Cleaned_data = pd.merge(Cleaned_data,open_cl_inv, on="Complaint Number", how = 'left')
            
# Updates the site_CMPLNT_TAT with new complaint numbers.
complaints_to_update = set(Cleaned_data["Complaint Number"])-set(site_CMPLNT_TAT_df["Complaint Number"]) 
wb_site_CMPLNT_TAT = openpyxl.load_workbook("./Reference/site_CMPLNT_TAT.xlsx")
sheet_TAT = wb_site_CMPLNT_TAT["Sheet1"]
for i,complaint in enumerate(complaints_to_update):
    maxrow = str(sheet_TAT.max_row+1)
    rowupdate = "E"+maxrow
    sheet_TAT[rowupdate] = complaint
wb_site_CMPLNT_TAT.save("./Reference/site_CMPLNT_TAT.xlsx")

print("Consider updating the TATs in the 'site_CMPLNT_TAT' workbook with these TATs: \n",
      Cleaned_data[["Complaint Number","Site TAT"]][Cleaned_data["TAT Source"] =="Complaint Task Report" ])

##############################################################
################  COMPLAINT FILE NAME INPUT ##################
##############################################################

filename = input("Input Complaint Report File name ")
if filename == "":
    filename = "Complaint Report"

    
report = openpyxl.Workbook()

def CopyPaste_sheet(df,sheet):
    '''Function for copying and pasting sheets'''
    for i,v in enumerate(df.columns):
        sheet["A1:ZZ1"][0][i].value = v
        for j, d in enumerate(df.iloc[0:,i]):
            try:
                j=str(j+2)
                sheet["A"+j:"ZZ"+j][0][i].value = d
            except:
                pass
# Dashboard sheet
report.create_sheet(index = 0, title = "Dashboard")
dash_sheet = report.active
dash_sheet["A1"] = "Complaint Report Pulled: "
dash_sheet["B1"] = complaintpulled
dash_sheet["A2"] = "Regulatory Report Pulled: "
dash_sheet["B2"] = regpulled
dash_sheet["A3"] = "Complaint Task Report Pulled: "
dash_sheet["B3"] = cmplnt_task_pulled
dash_sheet.column_dimensions["A"].width = 30


# Open Complaints Sheet
report.create_sheet(index = 0, title = "Open Complaints")
open_complaint_sheet = report.active
CopyPaste_sheet(Open_complaint,open_complaint_sheet)
open_complaint_sheet.column_dimensions["A"].width = 25
open_complaint_sheet.column_dimensions["B"].width = 25
open_complaint_sheet.column_dimensions["C"].width = 15
open_complaint_sheet.column_dimensions["D"].width = 20
open_complaint_sheet.column_dimensions["E"].width = 40
open_complaint_sheet.column_dimensions["F"].width = 22
open_complaint_sheet.column_dimensions["G"].width = 40
open_complaint_sheet.column_dimensions["H"].width = 40
open_complaint_sheet.column_dimensions["I"].width = 15
open_complaint_sheet.column_dimensions["J"].width = 15
open_complaint_sheet.column_dimensions["K"].width = 15
open_complaint_sheet.freeze_panes = "C2"

# Cleaned data sheets
report.create_sheet(index = 0, title = "Cleaned Data")
cleaned_data_sheet = report.active
CopyPaste_sheet(Cleaned_data,cleaned_data_sheet)
cleaned_data_sheet.column_dimensions["A"].width = 25
cleaned_data_sheet.column_dimensions["B"].width = 25
cleaned_data_sheet.column_dimensions["C"].width = 15
cleaned_data_sheet.column_dimensions["D"].width = 20
cleaned_data_sheet.column_dimensions["E"].width = 40
cleaned_data_sheet.column_dimensions["F"].width = 22
cleaned_data_sheet.column_dimensions["G"].width = 40
cleaned_data_sheet.column_dimensions["H"].width = 40
cleaned_data_sheet.column_dimensions["I"].width = 15
cleaned_data_sheet.column_dimensions["J"].width = 15
cleaned_data_sheet.column_dimensions["K"].width = 15
cleaned_data_sheet.freeze_panes = "A2"
cleaned_data_sheet.freeze_panes = "C2"


# Trending Sheet
report.create_sheet(index = 0, title = "Trending")
trend_data_sheet = report.active
CopyPaste_sheet(Trend_data,trend_data_sheet)
trend_data_sheet.column_dimensions["A"].width = 15
trend_data_sheet.column_dimensions["B"].width = 25
trend_data_sheet.column_dimensions["C"].width = 25
trend_data_sheet.column_dimensions["D"].width = 25
trend_data_sheet.column_dimensions["E"].width = 20
trend_data_sheet.column_dimensions["F"].width = 40
trend_data_sheet.column_dimensions["G"].width = 20
trend_data_sheet.freeze_panes = "A2"
trend_data_sheet.freeze_panes = "C2"

#Raw data sheet
report.create_sheet(index = 0, title = "Raw Data")
raw_data_sheet = report.active
CopyPaste_sheet(raw_data,raw_data_sheet)
raw_data_sheet.freeze_panes = "A2"
raw_data_sheet.freeze_panes = "C2"

report.save('./Complaints/'+filename+'.xlsx')

##################################################
##############  OPEN FILE INPUT ##################
##################################################
openfile_question = input("Open file? y/n ")
perform_trending = input("Perform trending? y/n ")
if openfile_question == 'y':
    specialfilename = "\\Complaints\\"+ filename+'.xlsx'
    filepath = rootpath+specialfilename
    os.startfile(filepath)

##################################################
##############  Complaint Trending ###############
##################################################




if perform_trending == 'n':
    exit()
else:
    pass
def trending(PF,Start,End):
    trend_data = data[(data["Product Family"]==PF)
    & (data["Complaint Create Date"]>Start)&(data["Complaint Create Date"]<End)]
    trend_data = trend_data[['Complaint Create Date','Complaint Number','Product','Lot # / Work Order',
                 'Investigated Failure Mode','Investigation Summary','Risk',"Product Family"]]
    return(trend_data)
def CopyPaste_histo_data(df,sheet):
    '''Function for copying and pasting sheets'''
    for i,v in enumerate(df.columns):
        sheet["h1:AQ1"][0][i].value = v
        for j, d in enumerate(df.iloc[0:,i]):
            try:
                j=str(j+2)
                sheet["H"+j:"AP"+j][0][i].value = d
            except:
                pass
Syringe_pf = ['1','2', '3', 'Sub', '7', '8', '5', '412', 'Tips', 'Unk NS']
Containers_pf = ['G1','G2', 'G3', 'G5', 'G7', 'G8','G10', 'G12', 'G18','HG', 'AS',
       'CT', 'GG', 'MS', 'SS',]

Start = input("enter start date for trending (MM/DD/YEAR)")
End = input("enter end date for trending (MM/DD/YEAR)")
if Start == "":
    Start ="10/01/2019"
if End == "":
    End = "9/30/2020"

#Start,End = "10/01/2019", "9/30/2020"


filename = input("Enter file name for CFF Trending: ")
if filename == "":
    filename = "CFF Complaint Trending"

CFF_CIP_trending = openpyxl.Workbook()

# missing lids sheet
missing_lids = data[['Complaint Create Date','Complaint Number','Product','Lot # / Work Order',
                 'Investigated Failure Mode','Risk',"Product Family"]][
    (data['Investigated Failure Mode']=="MISSING LIDS")
    &(data["Complaint Create Date"]>Start)
    &(data["Complaint Create Date"]<End)]
CFF_CIP_trending.create_sheet(index = 0, title = "Missing_Lids")
# need to generate data for histogram here.
pf = []
counts = []
for product_family in missing_lids["Product Family"].unique():
    pf.append(product_family)
    counts.append(missing_lids["Product Family"][missing_lids["Product Family"]==product_family].count())
results = {"Product Family":pf,"Count of Missing Lid Complaints":counts}
missing_lid_pf_df = pd.DataFrame(results)
missing_lid_pf_df = missing_lid_pf_df.sort_values(by = "Count of Missing Lid Complaints",ascending = False)

sheet = CFF_CIP_trending.active
CopyPaste_sheet(missing_lids,sheet)
CopyPaste_histo_data(missing_lid_pf_df,sheet)

chart1 = BarChart()
chart1.type = 'col'
chart1.style = 4
chart1.title = "Missing Lids"
chart1.grouping = "clustered"
chart1.y_axis.title = "Number of Complaints"
chart1.x_axis.title ="Problem Codes"
chartdata = Reference(sheet, min_col=9, min_row=1, max_row=len(missing_lid_pf_df)+1, max_col=9)
cats = Reference(sheet, min_col=8, min_row=2, max_row=len(missing_lid_pf_df)+1,max_col=8)
chart1.add_data(chartdata, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
sheet.add_chart(chart1, "M2")


# Product Family Sheets
CIP_Trending=[]
for pfs in Containers_pf:
    test = trending(pfs,Start,End)
    CIP_Trending.append(test)
for i, pf in enumerate(Containers_pf): 
    CFF_CIP_trending.create_sheet(index = 0, title = Containers_pf[i])
    sheet = CFF_CIP_trending.active
    CopyPaste_sheet(CIP_Trending[i][CIP_Trending[i]["Investigated Failure Mode"]!="MISSING LIDS"],sheet)

    #generates the data for the histogram
    fm = []
    counts = []
    risk_trending = []
    for ifm in CIP_Trending[i]["Investigated Failure Mode"].unique():
        if ifm == "MISSING LIDS":
            pass
        else:
            try: # Skips indexerrors for nans
                risk_trending.append(risk["Risk"][risk["Failure Mode"]==ifm].iloc[0])
                fm.append(ifm)
                counts.append(CIP_Trending[i]["Investigated Failure Mode"][(CIP_Trending[i]["Investigated Failure Mode"]==ifm)].count())
            except IndexError:
                if str(ifm) != 'nan':
                    print(ifm+' Failure Mode not found in risk reference document')
                else:
                    pass
    results = {"Investigated Failure Mode":fm,"Risk":risk_trending,"Count of Failure Mode":counts}
    histogram_data_df = pd.DataFrame(results)
    histogram_data_df = histogram_data_df.sort_values(by = "Count of Failure Mode",ascending = False)

    sheet = CFF_CIP_trending[Containers_pf[i]]

    CopyPaste_histo_data(histogram_data_df,sheet)


    chart1 = BarChart()
    chart1.type = 'col'
    chart1.style = 4
    chart1.title = Containers_pf[i]
    chart1.grouping = "clustered"
    chart1.y_axis.title = "Number of Complaints"
    chart1.x_axis.title ="Problem Codes"
    chartdata = Reference(sheet, min_col=10, min_row=1, max_row=len(histogram_data_df)+1, max_col=10)
    cats = Reference(sheet, min_col=8, min_row=2, max_row=len(histogram_data_df)+1,max_col=9)
    chart1.add_data(chartdata, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    sheet.add_chart(chart1, "M2")
    CFF_CIP_trending.save('./Complaints/'+filename+'.xlsx')

#Raw data sheet
CFF_CIP_trending.create_sheet(index = 0, title = "Raw Data")
raw_data_sheet = CFF_CIP_trending.active
CopyPaste_sheet(raw_data,raw_data_sheet)
raw_data_sheet.freeze_panes = "A2"
raw_data_sheet.freeze_panes = "C2"
CFF_CIP_trending.save('./Complaints/'+filename+'.xlsx')

filename = input("Enter file name for SFF Trending: ")
if filename =="":
    filename = "SFF CIP Complaint Trending"
SFF_CIP_trending = openpyxl.Workbook()
# Product Family Sheets
CIP_Trending=[]
for pfs in Syringe_pf:
    test = trending(pfs,Start,End)
    CIP_Trending.append(test)
for i, pf in enumerate(Syringe_pf): 
    SFF_CIP_trending.create_sheet(index = 0, title = Syringe_pf[i])
    sheet = SFF_CIP_trending.active
    CopyPaste_sheet(CIP_Trending[i],sheet)

    #generates the data for the histogram
    fm = []
    counts = []
    risk_trending = []
    for ifm in CIP_Trending[i]["Investigated Failure Mode"].unique():

        risk_trending.append(risk["Risk"][risk["Failure Mode"]==ifm].iloc[0])
        fm.append(ifm)
        counts.append(CIP_Trending[i]["Investigated Failure Mode"][(CIP_Trending[i]["Investigated Failure Mode"]==ifm)].count())
    results = {"Investigated Failure Mode":fm,"Risk":risk_trending,"Count of Failure Mode":counts}
    histogram_data_df = pd.DataFrame(results)
    histogram_data_df = histogram_data_df.sort_values(by = "Count of Failure Mode",ascending = False)
    sheet = SFF_CIP_trending[Syringe_pf[i]]

    CopyPaste_histo_data(histogram_data_df,sheet)


    chart1 = BarChart()
    chart1.type = 'col'
    chart1.style = 4
    chart1.title = Syringe_pf[i]
    chart1.grouping = "clustered"
    chart1.y_axis.title = "Number of Complaints"
    chart1.x_axis.title ="Problem Codes"
    chartdata = Reference(sheet, min_col=10, min_row=1, max_row=len(histogram_data_df)+1, max_col=10)
    cats = Reference(sheet, min_col=8, min_row=2, max_row=len(histogram_data_df)+1,max_col=9)
    chart1.add_data(chartdata, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    sheet.add_chart(chart1, "M2")
    SFF_CIP_trending.save('./Complaints/'+filename+'.xlsx')

SFF_CIP_trending.create_sheet(index = 0, title = "Raw Data")
raw_data_sheet = SFF_CIP_trending.active
CopyPaste_sheet(raw_data,raw_data_sheet)
raw_data_sheet.freeze_panes = "A2"
raw_data_sheet.freeze_panes = "C2"
SFF_CIP_trending.save('./Complaints/'+filename+'.xlsx')
