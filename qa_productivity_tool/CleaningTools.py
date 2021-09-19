


def mostrecentreport(report,path=r"C:\Users\steven.hansen01\Downloads"):
    """ Enter the name of report (full report name not necessary) and the path
        if not using the Downloads folder.
        Returns the filename it found, the path, and the date the file was created.
        For example
        filename,path,date = MMR.mostrecentreport('NC Report - MRB')
        or
        """
    import os
    import pandas as pd
    entries = {}
    #path = r"C:\Users\steven.hansen01\Downloads"
    with os.scandir(path) as it:
        for entry in it:
            if entry.name.startswith(report):
            #print(entry.name)
                entries.update({entry.name:[entry.stat().st_ctime,entry.path]})
    for key in entries:
        if entries[key] == max(entries.values()):
            timepulled = pd.Timestamp(entries[key][0], unit = 's')
            day,month,year = timepulled.day,timepulled.month,timepulled.year
            datepulled = str(month)+"/"+str(day)+"/"+str(year)
            return key,entries[key][1], datepulled

def folder_creator(to_path,foldername):
        time_stamp = time.strftime('_%m_%d_%Y_%H%M%S',time.localtime(time.time()))
        created_path = to_path + foldername + time_stamp
        os.mkdir(created_path)
        r_sub_folder_path = created_path + r"\\original_data"
        os.mkdir(r_sub_folder_path)
        fig_sub_folder_path = created_path + r"\\figures"
        os.mkdir(fig_sub_folder_path)
        return created_path, r_sub_folder_path, fig_sub_folder_path

##############################################################
######################  Report Cleaners    ###################
##############################################################

def tmpecc_remover(product_col):
    '''Removes "~TMPECC" from product codes
    example: data["Product"] = tmpecc_remover(data['Product'])'''
    def remover(pc):
        try:
            return pc.split('~')[0]
        except:
            if str(pc) == "None" or str(pc) == "nan":
                pass
            else:
                print('Failed for ',pc)
    return list(map(remover,product_col))

def index_compiler(indexer, df,retain_duplicates=[]):
    ''' For situations when we want to use an indexer that is listed
    in multiple rows because of every row represents a unique column.  
    For example NC# is the desired index but there are duiplciate NC
    rows because each row represents disposition or task activities.
    retain_duplicates takes a list of columns and does not remove duplicates
    for that column'''
    import pandas as pd
    index = df[indexer].unique()
    cols = df.columns # columns not including indexer
    new_data = pd.DataFrame(columns = cols)
    new_data[indexer] = index
    for i, ind in enumerate(index): 
        # cycles through unique indexes
        new_cols = cols[cols != indexer]
        for c in new_cols:
            if len(df[c][df[indexer]==ind].unique())>1:
                if c in retain_duplicates:
                    new_data.loc[i,c] = str(list(df[c][df[indexer]==ind].dropna())).strip('[]').strip("'")
                else:
                    new_data.loc[i,c] = str(list(df[c][df[indexer]==ind].dropna().unique())).strip('[]').strip("'")
            elif str(ind) == 'nan':
                pass
            else:
                new_data.loc[i,c] = df[c][df[indexer]==ind].unique()[0]
    new_data = new_data.reindex()
    return new_data

def date_converter(df,col_name_list):
    ''' Takes the names of columns that are in the default SmartSolve
    date format and converts it to excel friendly date format'''
    import pandas as pd
    for col in col_name_list:
        df[col] = df[col].fillna(0)
        df[col]=pd.to_datetime(df[col],errors = 'coerce')
        df[col] = df[col].dt.date


def phase_compiler(df):
    ''' index_compiler doesn't work well for phases. this is to amend that'''
    task_type = ["Verify Non Conformance Initiation", 'Adhoc','Approve Disposition Plan',
    'Approve Investigation Level 1','Containment Action', 'Execute Disposition',
    'Investigate Non Conformance', 'Plan Disposition','Sign-off Action Plans','Verify Non Conformance Closure']
    import pandas as pd
    import numpy as np
    pd.options.mode.chained_assignment = None
    # creates columns
    for task in task_type:
        df[task] = np.nan
    for nc in df['NC Number'].unique():
        for taskname in df['Task Name'][df['NC Number']==nc].unique():#cycles through tasks assinged for NC
            for task in task_type:
                if taskname == task:
                    if str(df["Task Owner"][df['Task Name']==task]) == 'nan':
                        pass
                    else:
                        
                        df[task][df['NC Number']==nc] = df["Task Owner"][df['Task Name']==task]
                    #df[task][df['NC Number']==nc] = df['Task Owner'][(df['Task Name']==task) & (df['NC Number']==nc)]
                else:
                    pass
    return df 

def task_counter(df):
    import pandas as pd
    ''' Depreciated see task_counterII
    gives a df with a coung how many NCs are waiting on'''
    task_type = ["Verify Non Conformance Initiation", 'Adhoc','Approve Disposition Plan',
    'Approve Investigation Level 1','Containment Action', 'Execute Disposition',
    'Investigate Non Conformance', 'Plan Disposition','Sign-off Action Plans','Verify Non Conformance Closure']
    for i,age in enumerate(df["Age of NC"]):
        if age >= 60:
            df.loc[i,'Over 60'] = 1
        else:
            df.loc[i,'Over 60'] = 0
    task_count = pd.DataFrame(columns = task_type)
    for i,task in enumerate(task_type):
        task_count.loc['Over 60 Days',task] = df['Task Name'][(df['Task Name'] == task) & (df['Over 60']>0)].count()
    for task in task_type:
        task_count.loc['Under 60 Days',task] = df['Task Name'][(df['Task Name'] == task) & (df['Over 60']==0)].count()
    return task_count

def task_counterII(df):
    import pandas as pd
    tasks = {"Verify Non Conformance Initiation":0, 'Adhoc':0,'Approve Disposition Plan':0,
    'Approve Investigation Level 1':0,'Containment Action':0, 'Execute Disposition':0,
    'Investigate Non Conformance':0, 'Plan Disposition':0,'Sign-off Action Plans':0,'Verify Non Conformance Closure':0}
    Over60 = {"Verify Non Conformance Initiation":0, 'Adhoc':0,'Approve Disposition Plan':0,
    'Approve Investigation Level 1':0,'Containment Action':0, 'Execute Disposition':0,
    'Investigate Non Conformance':0, 'Plan Disposition':0,'Sign-off Action Plans':0,'Verify Non Conformance Closure':0}
    Under60 = {"Verify Non Conformance Initiation":0, 'Adhoc':0,'Approve Disposition Plan':0,
    'Approve Investigation Level 1':0,'Containment Action':0, 'Execute Disposition':0,
    'Investigate Non Conformance':0, 'Plan Disposition':0,'Sign-off Action Plans':0,'Verify Non Conformance Closure':0}
    mod_df = df
    for i, task_name in enumerate(mod_df['Task Name']):
        if str(task_name) == 'nan':
            break
        task_name = task_name.strip("[]'")
        task_name = task_name.split(',')
        if mod_df.loc[i,'Over 60 days']== 1: 
            for t in task_name:
                t = t.strip(" '")
                Over60[t] +=1
        else:
            for t in task_name:
                t = t.strip(" '")
                Under60[t] +=1
    list_of_dics = [Over60, Under60]
    tasks_df = pd.DataFrame(list_of_dics).transpose()
    tasks_df = tasks_df.rename(columns={0:'Over 60 Days',1:'Under 60 Days'})
    return tasks_df

    


def cmplnt_report_cleaner(cmplnt_report):
    # remove unused or unnessecary columns
    allna = ["Product Group","Reported Component", "Reported Component Name","Reported Vendor Name"]
    useless_columns = ["Note Type", "Sales Rep","Sales Support","Customer Service"]
    columns_low_use = ["Manufactured Date","Investigated Component","Investigated Component Name","Investigated Component Lot Number",
                  "Investigated Mfg. SKU Number","Investigated Vendor Name"]
    columns_to_drop = allna+useless_columns+columns_low_use
    cmplnt_report = cmplnt_report.drop(columns_to_drop,axis=1)
    cmplnt_report["Product"] = list(map(TMPECC_remover,cmplnt_report['Product']))
    
    return cmplnt_report
    
##############################################################
#### ###################   Excel Tools    ################
##############################################################

def copy_paste_sheet(df,sheet):
    import openpyxl
    '''Function for copying and pasting sheets'''
    for i,v in enumerate(df.columns):
        sheet["A1:ZZ1"][0][i].value = v
        for j, d in enumerate(df.iloc[0:,i]):
            try:
                j=str(j+2)
                sheet["A"+j:"ZZ"+j][0][i].value = d
            except:
                pass


def excel_bar_chart(ref_sheet, dis_sheet,title, y_axis_title, x_axis_title,y_col_num,x_col_num,cell):
    '''Generate a bar chart from excel data.
    col_num is the number of the column of the data (i.e. col "A" is 1).
    ref_sheet is the sheet where the data is referenced and
    dis_sheet is the sheet where the chart will be copied to
    '''
    from openpyxl.chart import BarChart, Series, Reference
    chart = BarChart()
    chart.type = 'col'
    chart.style = 4
    chart.title = title
    chart.grouping = "clustered"
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    data = Reference(ref_sheet, min_col = y_col_num,max_col = y_col_num,min_row = 1,max_row=ref_sheet.max_row)
    #setting max_row to 100.  Trying not to require dataframe
    cats = Reference(ref_sheet,min_col = x_col_num,max_col = x_col_num,min_row =2,max_row=ref_sheet.max_row)
    # leaving max_row empty
    chart.add_data(data,titles_from_data = True)
    chart.set_categories(cats)
    chart.shape = 3
    dis_sheet.add_chart(chart,cell)

def excel_col_width(sheet,cols_width):
    '''Takes in a dict of col letters and widths and outputs the column width'''
    import openpyxl
    for col in cols_width:
        sheet.column_dimensions[col].width = cols_width[col]

def excel_text_wrap(sheet,col_interval):
    '''Text wraps all columns in interval (i.e. if goal is to text wrap all columns
    from A to O, enter the col_interval as [1,15]'''
    ### not working for intevals [1,31] or [1,30] not sure why.
    # considering a different approach - Try selecting columns and applying
    # text wrap by column
    from openpyxl.styles import Alignment
    for row in sheet.iter_rows(min_col = col_interval[0],max_col=col_interval[1],min_row= sheet.min_row,max_row = sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)



