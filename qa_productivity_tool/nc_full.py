import pandas as pd
import numpy as np
import openpyxl
import time
from datetime import datetime,timedelta
import datetime as dt
import matplotlib.pyplot as plt

import os
plt.style.use('fivethirtyeight')
['Solarize_Light2', '_classic_test_patch', 'bmh', 'classic', 'dark_background', 'fast', 'fivethirtyeight', 'ggplot', 'grayscale', 'seaborn', 'seaborn-bright', 
'seaborn-colorblind', 'seaborn-dark', 'seaborn-dark-palette', 'seaborn-darkgrid', 'seaborn-deep', 'seaborn-muted', 'seaborn-notebook', 'seaborn-paper', 'seaborn-pastel', 'seaborn-poster', 'seaborn-talk', 'seaborn-ticks', 'seaborn-white', 'seaborn-whitegrid', 'tableau-colorblind10']

from .quality_report import Quality_Report
from .CleaningTools import *
from .Reference import *



def test_clipboard():
    nc_full = NC_Full()
    #nc_full.clipboard_data('clean')
    nc_full.clipboard_data('clean')
def test_default_report():
    nc_full = NC_Full()
    filename = 'NC_Full_test' + str(time.time()).split('.')[0]
    nc_full.default_report(filename,openfile='y')
def test_metrics():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    yrTAT,included_nc_TAT, thirtyTAT,fyTAT = nc_full.metrics()
    print(yrTAT,included_nc_TAT,thirtyTAT,fyTAT) 
def test_nc_closed_month_td():
    nc_full = NC_Full()
    results = nc_full.nc_closed_month_td(2.0,2021.0)
    results.to_clipboard()
def test_run_report():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    data = nc_full.run_report()
    data.to_clipboard()   
def test_visual_root_cause_loc():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    cff,sff = nc_full.visual_root_cause_loc('1/01/21')
    cff.to_clipboard()
def test_visual_open_closed():
    nc_full = NC_Full()
    nc_full.visual_open_closed(value_stream='Container')
def test_analysis_missing_lids():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    nc_full.missing_lid_ncs().to_clipboard()
def test_visual_open_closed():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    nc_full.visual_open_closed()
    nc_full.visual_open_closed(value_stream='Container')
    nc_full.visual_open_closed(value_stream='Syringe')
    nc_full.visual_open_closed(value_stream='Non-FF')

def test_num_open_timeline():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    nc_full.run_report()
    nc_full.num_open_timeline()
    nc_full.num_open_timeline(value_stream='Syringe')
    nc_full.num_open_timeline(value_stream='Container')
    nc_full.num_open_timeline(value_stream='Non-FF')
def test_sub_num_open_timeline():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    nc_full.run_report()
    nc_full.sub_num_open_timeline(from_date=dt.date(2021,6,1))
def test_visual_opened_ncs():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    nc_full.run_report()
    nc_full.visual_opened_ncs(date_i = dt.date(2019,7,1),
    split_date = dt.date(2020,6,30),end_date=dt.date(2021,7,1),  ff = "CL")
def test_tat_30day_algo():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    nc_full.run_report()
    nc_full.nc_hist_open_tat().to_clipboard()
    nc_full.visual_waterfallish()
def test_tat_fy_algo():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    df = nc_full.run_report()
    print(nc_full._tat_fy_algo())
def test_check_for_lot_number():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    print(nc_full.check_for_lot_number('18M17263'))
def test_root_cause_loc():
    nc_full = NC_Full()
    nc_full.mostrecentreport()
    nc_full.test_root_cause_loc(start_date = '01/01/2021')


protocol = test_tat_fy_algo

class NC_Full(Quality_Report):
    ''' Cleans the report 'NC Full Report' and provides dataframes, basic excel report(s) and visualizations'''
    def __init__(self):
        super().__init__(report_name="NC Full Report")
        #self.r_filename, self.path, self.date = ct.mostrecentreport("NC Full Report")
        self.status_of_report = 0
    #def meta_data(self): # obsoleted by subclassing
    #    return self.r_filename, self.path, self.date
    def metrics(self):
        if self.status_of_report == 0:
            self.run_report()
        self.closed_ncs = self.clean_data_df[self.clean_data_df["Status"]=="CLOSED"]
        # Yr TAT metric
        self.yr_ago = pd.to_datetime(datetime.today() - timedelta(days = 365))
        self.ncs_closed_yr = self.closed_ncs[pd.to_datetime(self.closed_ncs['Closed Date'])>=self.yr_ago]
        self.yrTAT = sum(self.ncs_closed_yr["TAT (Days)"].astype('float'))/len(self.ncs_closed_yr["TAT (Days)"])
        # Yr TAT metric OLD -  excluding old calibration NCs
        self.excluded_nc = pd.read_excel(open(r'C:\\Users\steven.hansen01\data_automation\\venv\\excluded_ncs.xlsx','rb'))
        self.excluded_nc = pd.merge(self.ncs_closed_yr,self.excluded_nc,on = 'NC Number',how='left')
        self.included_nc = self.excluded_nc[(self.excluded_nc['Exclude']!='y') & (self.excluded_nc['Exclude']!='Old Calibration') & (self.excluded_nc['Exclude']!='Incoming')]
        self.included_nc_TAT = sum(self.included_nc["TAT (Days)"].astype('float'))/len(self.included_nc["TAT (Days)"])
        # Fiscal Year TAT metric
        self.ncs_closed_FY22 = self.closed_ncs[pd.to_datetime(self.closed_ncs['Closed Date'])>=dt.datetime(2021,7,1)]
        self.fy22_TAT = sum(self.ncs_closed_FY22["TAT (Days)"].astype('float'))/len(self.ncs_closed_FY22["TAT (Days)"])


        # 30 day TAT Metric
        self.thirty_days_ago = pd.to_datetime(datetime.today() - timedelta(days = 30))
        self.ncs_closed_30days = self.closed_ncs[pd.to_datetime(self.closed_ncs['Closed Date'])>=self.thirty_days_ago]
        self.thirtyTAT = sum(self.ncs_closed_30days["TAT (Days)"].astype('float'))/len(self.ncs_closed_30days["TAT (Days)"])

        return (self.yrTAT,self.included_nc_TAT,self.thirtyTAT,self.fy22_TAT)       
    def clipboard_data(self,type_of_data = 'clean'):
        if self.status_of_report == 0:
            self.run_report()
        if type_of_data == 'clean':    
            self.clean_data_df.to_clipboard()
            print('Check Clipboard')
        if (type_of_data == 'original') or (type_of_data == 'raw'):
            self.r_data.to_clipboard()
            print('Check Clipboard')
        if type_of_data == 'metrics':
            self.metrics()
            self.ncs_closed_30days.to_clipboard()
            print('Check Clipboard')
    def run_report(self):
        ''' No frills just the data from the report organized better'''
        #self.mostrecentreport()
        if self.status_of_report == 1:
            return self.clean_data_df
        print('Running NC Full Report...')
        self.r_data = pd.read_excel(open(self.path,'rb'))
        self.data = self.r_data.iloc[4:]
        self.data = self.data.rename(columns = self.data.iloc[0,:])
        self.data = self.data.drop(4)
        date_fields = ['Created Date','Report Date','Occurrence Date','Closure Due Date','Closed Date']
        #ct.date_converter(self.data,date_fields)
        
        self.data['Product'] = list(map(self.tmpecc_remover,self.data['Product']))
        self.data = self.data[self.data["NC Number"] != self.data.iloc[-1,0]]
        self.data["TAT (Days)"] = (self.data['Closed Date'] - self.data['Created Date']).dt.days

        self.data = self.date_converter(self.data,date_fields)
        print('Compiling index, sit tight...')
        #self.data = ct.index_compiler('NC Number',self.data)
        self.data = self.index_compressor(self.data,'NC Number')
        
        # filter out failure modes
        self.data = self.data[(self.data['Failure Mode'] !='PR-Planned Deviation') & 
        (self.data['Failure Mode']!='Validation') & (self.data['Failure Mode']!='Validation - Process Not to Protocol Requirements') ]

        self.clean_data_df = self.data.reset_index(drop=True)
        # product family / value stream
        reference = Reference()
        self.clean_data_df = reference.evaluate_pc(self.clean_data_df)
        
        self.clean_data_df = self.clean_data_df.reset_index(drop=True)
        list_of_cols_to_keep = ['NC Number','Created Date','Created By', 'NC Source',
        'NC Type','Discovery/Plant Area','Discovery/Plant Area Name','Item Type','Other Item','Product','Product Name','Initial Failure Mode','Description','Lot Number',
        'Quantity Affected','Immediate Actions','Report Date','Occurrence Date','Failure Mode',
        'Failure Mode Description','Root Cause','Root Cause Description','Disposition',
        'Dispositioned Quantity','Disposition Cost','Phase','Status','Closure Due Date',
        'Closed Date',"TAT (Days)",'Product Family','Value Stream']
        self.clean_data_df = self.clean_data_df[list_of_cols_to_keep]
        self.status_of_report = 1
        return self.clean_data_df
    def default_report(self,filename,openfile = 'n',to_path = r"C:\\Users\\steven.hansen01\\data_automation\\venv\\Test_Reports\\"):
        
        if self.status_of_report == 0:
            self.run_report()
        print('Building report...')
        self.filename = filename
        self.report = openpyxl.Workbook()
        # meta_data
        self.report.create_sheet(index = 0, title = "meta_data ")
        self.meta_sheet = self.report.active
        self.meta_sheet["A1"] = "Original Data Filename"
        self.meta_sheet["A2"] = self.r_filename
        self.meta_sheet["B1"] = "Data generated: "
        self.meta_sheet["B2"] = self.date
        # raw_data
        self.report.create_sheet(index=0,title = "raw_data")
        self.raw_data_ws = self.report.active
        ct.copy_paste_sheet(self.r_data,self.raw_data_ws)
        # clean_data
        self.report.create_sheet(index = 0,title="clean_data")
        self.clean_data_ws = self.report.active
        ct.copy_paste_sheet(self.clean_data_df,self.clean_data_ws)
        cols_width = {'A':16,'B':12.30,'C':23,'D':9,'E':9,'F':10,'G':15,'H':12,'I':12,'J':12,"K":12,"L":22,"M":25,"N":25,"O":25,
        'P':25,'Q':29,'R':12,'S':12,'T':14,'U':25,'V':20,'W':20,'X':10,'Y':20,'Z':25,
        'AA':10,'AB':15,'AC':10,'AD':16,'AE':16}
        ct.excel_col_width(self.clean_data_ws,cols_width)
        # wrapping text is not working at 30 or 31 columns tried doing it in
        # batches and still is not working.  It's causing a PermissionError,
        # which makes me think that it is tying some processing up downstream
        # or something.  
        ct.excel_text_wrap(self.clean_data_ws,[1,15])
        #ct.excel_text_wrap(self.clean_data_ws,[16,31])

        # save and open
        self.report.save(to_path+self.filename+'.xlsx')
        print('File saved in '+to_path)
        if openfile =='y':
            print('Opening file')
            os.startfile(to_path+self.filename+'.xlsx')
    def nc_closed_month_td(self,month,year):
        # this method is pending obsoletion 
        '''Used in the ncTAT metric in nc_reports.py 
        returns a df of all complaints closed in given month and year'''
        if self.status_of_report == 0:
            self.run_report()
        def to_month(d):
            return pd.to_datetime(d).month
        def to_year(d):
            return pd.to_datetime(d).year
        self.clean_data_df["Closed Month"] = list(map(to_month,self.clean_data_df['Closed Date']))
        self.clean_data_df["Closed Year"] = list(map(to_year,self.clean_data_df['Closed Date']))
        self.nc_closed_df = self.clean_data_df.dropna(subset = ['Closed Date'])
        self.nc_closed_df = self.nc_closed_df[(self.nc_closed_df["Closed Month"] == month) &
        (self.nc_closed_df['Closed Year'] == year)]

        return self.nc_closed_df[['NC Number','Created Date','Closed Date']]
    def visual_root_cause_loc(self,start_date,save_to_path = './Figures'):
        '''Creates bar chart and returns 1 df per value stream (cff, sff).  Future state
        might also save the data to the save_to_path as well '''
        self.run_report()
        df = self.clean_data_df[self.clean_data_df['Created Date'] > pd.to_datetime(start_date)]
        df_sff = df[df['Value Stream']=='Syringe']
        df_cff = df[df['Value Stream']=='Container']
        r_c_loc_sff = pd.DataFrame()
        for i,loc in enumerate(df_sff['Discovery/Plant Area'].unique()):
            if str(loc)=='None':
                loc = 'No Loc'
                r_c_loc_sff.loc[i,"Root Cause Location"] = loc
                r_c_loc_sff.loc[i,'Number of NCs'] = df_sff['Discovery/Plant Area'].isna().sum()
            else:
                r_c_loc_sff.loc[i,"Root Cause Location"] = loc
                r_c_loc_sff.loc[i,'Number of NCs'] = df_sff['NC Number'][df_sff['Discovery/Plant Area']==loc].count()
        r_c_loc_cff = pd.DataFrame()
        for i, loc in enumerate(df_cff['Discovery/Plant Area'].unique()):
            if str(loc)=='None':
                loc = 'No Loc'
                r_c_loc_cff.loc[i,"Root Cause Location"] = loc
                r_c_loc_cff.loc[i,'Number of NCs'] = df_cff['Discovery/Plant Area'].isna().sum()
            else:
                r_c_loc_cff.loc[i,"Root Cause Location"] = loc
                r_c_loc_cff.loc[i,'Number of NCs'] = df_cff['NC Number'][df_cff['Discovery/Plant Area']==loc].count()
        r_c_loc_sff = r_c_loc_sff.sort_values(by = 'Number of NCs',ascending = False)
        r_c_loc_cff = r_c_loc_cff.sort_values(by = 'Number of NCs',ascending = False)
        f, (axes) = plt.subplots(ncols=1,nrows=2,figsize = (20,10),sharey=True)
        axes[0].bar(x = r_c_loc_sff['Root Cause Location'],height = r_c_loc_sff['Number of NCs'])
        axes[0].set_title('SFF')
        axes[1].bar(x = r_c_loc_cff['Root Cause Location'],height = r_c_loc_cff['Number of NCs'])
        axes[1].set_title('CFF')
        f.suptitle("Root Cause Location by Value Stream",fontsize = 16)
        f.tight_layout()
        f.savefig(save_to_path+r'\\'+'Root Cause Location')

        plt.close(f)
        return df_cff, df_sff
    def test_root_cause_loc(self, start_date,save_to_path = './Figures'):
        self.run_report()
        root_cause_loc = {}
        for rc, frame in self.clean_data_df.groupby(by = 'Discovery/Plant Area'):
            root_cause_loc[rc] = len(frame)
        print(root_cause_loc)

    def visual_open_closed(self,value_stream = "CL",save_to_path = './Figures',yyyy=2020,mon=2,day=1 ):
        self.check_report_status()
        # if self.status_of_report == 0:
        #     self.run_report()
        if value_stream == 'CL':
            data = self.clean_data_df
            plot_title = "CL - Open_vs_Closed NCs"
        elif value_stream == 'Container':
            data = self.clean_data_df[self.clean_data_df['Value Stream']=='Container']
            plot_title = 'CL - Container Open vs Closed NCs'
        elif value_stream == 'Syringe':
            data = self.clean_data_df[self.clean_data_df['Value Stream']== 'Syringe']
            plot_title = 'CL - Syringe Open vs Closed NCs'
        elif value_stream == 'Non-FF':
            data = self.clean_data_df[(self.clean_data_df["Value Stream"]!="Syringe") & 
                        (self.clean_data_df["Value Stream"]!="Container")]
            plot_title = 'CL - Non-FF Open vs Closed NCs'
            

            
        value_to_check = pd.Timestamp(yyyy, mon, day)
        filter_mask = data['Created Date'] > value_to_check
        filtered_df = data[filter_mask]
        closed_filtered_mask = data["Closed Date"] > value_to_check
        closed_filtered_df = data[closed_filtered_mask]
        closed_filtered_df['Closed Date'] = pd.to_datetime(closed_filtered_df['Closed Date'],format = "%Y/%m/%d")
        filtered_df['Created Date'] = pd.to_datetime(filtered_df['Created Date'],format = "%Y/%m/%d")
        create_data = filtered_df["NC Number"].groupby([filtered_df['Created Date'].dt.year, filtered_df['Created Date'].dt.month]).count()
        create_data_df = pd.DataFrame(create_data)

        x = list(create_data_df.index)
        y = create_data_df['NC Number']
        openNCs = []
        for i, j in enumerate(create_data_df['NC Number']):
            openNCs.append(j)
        labels = []
        for d in list(create_data_df.index):
            labels.append(str(d))
        closed_data = closed_filtered_df["NC Number"].groupby([closed_filtered_df['Closed Date'].dt.year,
                                    closed_filtered_df['Closed Date'].dt.month]).count()
        closed_data_df = pd.DataFrame(closed_data)
        x_closed = list(closed_data_df.index)
        y_closed = closed_data_df
        closedNCs = []
        for ncs in closed_data_df["NC Number"]:
            closedNCs.append(ncs)
        closed_labels=[]
        for d in list(closed_data_df.index):
            closed_labels.append(str(d))
        plt.figure(figsize = (18,9))
        #fig = plt.plot(labels,openNCs,label = "Open NCs",color = 'b',linewidth = 8)
        plt.plot(labels,openNCs,label = "Open NCs",color = 'b',linewidth = 8)
        plt.plot(closed_labels,closedNCs,label = 'Closed NCs',color = 'g',linewidth = 8)
        plt.xlabel("Year Month",fontsize = 12)
        plt.ylabel("NCs",fontsize = 12)
        plt.legend(fontsize = 12)
        plt.tick_params(axis = 'x',labelsize=12)
        plt.tick_params(axis = 'y',labelsize=12)
        plt.xticks(rotation = 0)
        # plt.annotate(str(openNCs[-1])+" NCs Opened",xy=(labels[-1],openNCs[-1]),fontsize=9,
        #             xytext=(0.85, 0.5), textcoords='axes fraction',
        #             arrowprops=dict(facecolor='black', shrink=0.005),)
        # plt.annotate(str(closedNCs[-1])+" NCs Closed",xy=(labels[-1],closedNCs[-1]),fontsize=9,
        #             xytext=(0.8, 0.05), textcoords='axes fraction',
        #             arrowprops=dict(facecolor='black', shrink=0.005),)
        plt.tight_layout()
        #plt.title(plot_title,fontsize = 10)
        plt.suptitle(plot_title,fontsize = 12)
        plt.savefig(save_to_path+r'\\'+plot_title)
        
        #plt.show()
    
    def missing_lid_ncs(self,to_path = r"C:\\Users\\steven.hansen01\\data_automation\\venv\\Test_Reports\\"):
        self.check_report_status()
        data_df = self.clean_data_df[self.clean_data_df['Value Stream']=='Container']
        list_of_cols_to_keep= ['NC Number','Created Date','Discovery/Plant Area','Discovery/Plant Area Name',
        'Product','Product Name','Initial Failure Mode','Description','Lot Number','Failure Mode',
        'Failure Mode Description','Root Cause','Root Cause Description','Status','Closed Date',
        'Product Family','Value Stream']
        all_initial_fms = data_df['Initial Failure Mode'].unique()
        ml_fms = ['CFF-V-Quantity per Case','CFF-V-Packaging 1','CFF-V-Packaging 2']
        data_df = data_df[list_of_cols_to_keep][(data_df['Initial Failure Mode'] == 'CFF-V-Quantity per Case') | 
        (data_df['Initial Failure Mode'] == 'CFF-V-Packaging 1' ) | (data_df['Initial Failure Mode'] == 'CFF-V-Packaging 2') |
        (data_df['Initial Failure Mode'] == 'Packaging - Incorrect Count')]
        data_df = data_df[data_df['Description'].str.contains("lids")]
        return data_df
    def num_open_timeline(self,from_date = dt.date(2019,1,1),value_stream = 'CL',save_to_path = './Figures'):
        if (value_stream == 'Container') | (value_stream=='Containers'):
            data = self.clean_data_df[self.clean_data_df['Value Stream']=='Container']
        elif value_stream == 'Syringe':
            data = self.clean_data_df[self.clean_data_df['Value Stream']=='Syringe']
        elif value_stream == 'Non-FF':
            data = self.clean_data_df[(self.clean_data_df["Value Stream"]!="Syringe") & 
                        (self.clean_data_df["Value Stream"]!="Container")]
        else:
            data = self.clean_data_df
        num_open_dict = self._num_open_algo(from_date=from_date,data = data)
        plt.plot(num_open_dict.keys(),num_open_dict.values(),label = value_stream)
        plt.title("Number Open NCs in "+value_stream+" by Date")
        plt.xticks(rotation = 75)
        plt.tight_layout()
        plt.savefig(save_to_path+'\\'+"Open NCs in "+value_stream)
        plt.close()
        return num_open_dict
    def sub_num_open_timeline(self,from_date = dt.date(2019,1,1),save_to_path = './Figures',fig_title = 'Number of Open NCs by Date'):
        df = self.clean_data_df.copy()
        cff_df = df[df['Value Stream']=='Container']
        sff_df = df[df['Value Stream']=='Syringe']
        nonff_df = df[(df['Value Stream']!='Container')&
        (df['Value Stream']!='Syringe')]
        cff_dict = self._num_open_algo(data = cff_df, from_date=from_date)

        sff_dict = self._num_open_algo(data = sff_df,from_date=from_date)
        nonff_dict = self._num_open_algo(data = nonff_df,from_date =from_date)
        cl_dict = self._num_open_algo(data = self.clean_data_df.copy(),from_date = from_date)
        f,(axes) = plt.subplots(ncols=2,nrows=2,figsize = (20,10),sharey=True)
        xrotate = 30
        axes[0,0].plot(cff_dict.keys(),cff_dict.values())
        axes[0,0].set_title('CFF')
        axes[0,0].tick_params(labelrotation = xrotate)

        axes[1,0].plot(sff_dict.keys(),sff_dict.values())
        axes[1,0].set_title('SFF')
        axes[1,0].tick_params(labelrotation = xrotate)
        

        axes[0,1].plot(nonff_dict.keys(),nonff_dict.values())
        axes[0,1].set_title('Non-FF')
        axes[0,1].tick_params(labelrotation = xrotate)

        axes[1,1].plot(cl_dict.keys(),cl_dict.values())
        axes[1,1].set_title('All CL')
        axes[1,1].tick_params(labelrotation = xrotate)

        f.suptitle(fig_title,fontsize = 'x-large')
        f.tight_layout()
        f.savefig(save_to_path+'\\'+"Number of Open NCs by Date")
    def visual_disp_qty_vs_age(self):
        '''Scatter plot of disposition quantity vs age, to include closed and opened'''
        pass
    def convert_str_date(self,str_date):
            try:return dt.datetime.strptime(str_date,'%d-%b-%Y').date()
            except: pass
    def _num_open_algo(self,data,from_date = dt.date(2019,1,1)):
        to_date = dt.date.today()
        date_i = from_date
        num_open_dict = {}
        
        data['Created Date'] = list(map(self.convert_str_date, data['Created Date']))
        data['Closed Date'] = list(map(self.convert_str_date, data['Closed Date']))
        while date_i <= to_date:
            num_open = len(data[
                (data['Created Date']<=date_i) &
            ((data['Closed Date']>=date_i)|
            (data['Status']=='INWORKS'))])
            num_open_dict[date_i] = num_open
            date_i = date_i+dt.timedelta(days = 1)
        return num_open_dict
    def _tat_day_algo(self,daysback = 30,from_date = dt.date(2019,1,31)):
        ## using 1/31/2019 has the default since 1/1/2019 had no TATs
        to_date = dt.date.today()
        date_i = from_date
        date_tat = {}
        df = self.clean_data_df.copy()
        df['Created Date'] = list(map(self.convert_str_date, df['Created Date']))
        df['Closed Date'] = list(map(self.convert_str_date, df['Closed Date']))
        while date_i <= to_date:
            days_ago = pd.to_datetime(date_i - timedelta(days = daysback))
            ncs_in_scope = df[(df['Closed Date']>= days_ago) &
            (df['Closed Date']<=date_i) ]
            if len(ncs_in_scope['TAT (Days)'].astype('float')) != 0:
                date_tat[date_i] = np.nanmean(ncs_in_scope['TAT (Days)'].astype('float'))
            else:
                date_tat[date_i] = 0
            date_i = date_i+dt.timedelta(days = 1)
        return date_tat
    def _tat_fy_algo(self,from_date = dt.date(2021,7,1)):
        date_tat = {}
        date_i = from_date
        df = self.clean_data_df.copy()
        df['Created Date'] = list(map(self.convert_str_date, df['Created Date']))
        df['Closed Date'] = list(map(self.convert_str_date, df['Closed Date']))
        while date_i <= dt.date.today():
            days_ago = pd.to_datetime(date_i - (date_i - from_date))
            ncs_in_scope = df[(df['Closed Date']>=days_ago) &
            (df['Closed Date']<=date_i)]
            if len(ncs_in_scope['TAT (Days)'].astype('float')) != 0:
                date_tat[date_i] = np.nanmean(ncs_in_scope['TAT (Days)'].astype('float'))
            else:
                date_tat[date_i] = 0
            date_i = date_i+dt.timedelta(days = 1)
        return date_tat


        return date_tat
    def nc_hist_open_tat(self):
        num_open_dict = self._num_open_algo(data = self.clean_data_df.copy())
        num_open_df = pd.DataFrame(list(num_open_dict.values()),list(num_open_dict.keys()),columns = ['Open NCs'])
        last_30day_tat = self._tat_day_algo()
        last_30day_tat_df = pd.DataFrame(list(last_30day_tat.values()),list(last_30day_tat.keys()),columns = ['Monthly TAT'])
        last_yr_tat = self._tat_day_algo(daysback=365)
        last_yr_tat_df = pd.DataFrame(list(last_yr_tat.values()),list(last_yr_tat.keys()),columns = ['Yearly TAT'])
        fy_yr_tat = self._tat_fy_algo()
        fy_yr_tat_df = pd.DataFrame(list(fy_yr_tat.values()),list(fy_yr_tat.keys()),columns = ['FY TAT'])
        nc_hist_df = last_30day_tat_df.merge(last_yr_tat_df,left_index = True,right_index = True)
        nc_hist_df = num_open_df.merge(nc_hist_df,left_index=True,right_index = True)
        nc_hist_df = fy_yr_tat_df.merge(nc_hist_df,how = 'outer',left_index=True,right_index=True)
        self.nc_hist_df = nc_hist_df.reset_index().rename(columns={'index':'Date'})
        for i, _ in enumerate(self.nc_hist_df['Open NCs']):
            if i == 0:
                continue

            self.nc_hist_df.loc[i,'Change in NCs'] = self.nc_hist_df.loc[i,'Open NCs'] - self.nc_hist_df.loc[i-1,'Open NCs']
        return self.nc_hist_df
    def visual_waterfallish(self,save_to_path = "./Figures",days_going_back = 14 ):
        self.nc_hist_open_tat()
        plot_data = self.nc_hist_df.dropna()
        plot_data['Date'] = pd.to_datetime(plot_data['Date'])
        date_days_ago = dt.datetime.now()- dt.timedelta(days = days_going_back) 
        plot_data = plot_data[plot_data['Date']>date_days_ago]
        plot_data = plot_data.sort_values('Date')
        plot_data.to_clipboard()
        net = sum(plot_data['Change in NCs'].dropna())
        if net <0:
            net_color = 'g'
        else:
            net_color = 'r'
        x = plot_data['Date'].dt.strftime('%m-%d')
        y = plot_data['Change in NCs']
        colors = []
        for i in y:
            if i>0:
                colors.append('r')
            if i<=0:
                colors.append('g')
        f, (axes) = plt.subplots(ncols=2,nrows=1,figsize = (10, 5),sharey = True,
        gridspec_kw={'width_ratios': [1, .05]})
        axes[0].bar(x,y,color = colors)
        axes[0].axhline(0,linestyle='-',color = 'k')
        #axes[0].set_xlabel(rotation = 70)
        axes[0].tick_params(labelrotation=45)
        axes[1].bar(x = ['Net Change'],height = net,color = net_color)
        axes[1].axhline(0,linestyle='-',color = 'k')
        f.suptitle("Change in Open NCs Last "+str(days_going_back)+ " Days")
        f.tight_layout()
        
        #plt.show()
        f.savefig(save_to_path+r'\\'+"Waterfallish")
        plt.close(f)
    def visual_opened_ncs(self,date_i = dt.date(2019,8,1),
    split_date = dt.date(2020,5,1),end_date = dt.date(2021,4,1),ff = 'Syringe',save_to_path = './Figures',):
        from bokeh.plotting import figure, show
        from bokeh.models import ColumnDataSource
        from dateutil.relativedelta import relativedelta
        self.run_report()
        if ff != 'CL':df = self.clean_data_df[self.clean_data_df["Value Stream"]==ff]
        else: df = self.clean_data_df
        to_date = dt.date.today()
        pre_fg_inspection = {}
        post_fg_inspection = {}
        df["month"] = pd.to_datetime(df['Created Date']).dt.month
        df["year"] = pd.to_datetime(df['Created Date']).dt.year
        df.to_clipboard()
        while date_i <= to_date:
            num_opened = sum((df['month']==date_i.month) & (df['year']==date_i.year))
            if date_i > end_date:
                break
            elif date_i <= split_date:
                pre_fg_inspection[str(date_i.month)+'-'+str(date_i.year)] = num_opened
            else:
                post_fg_inspection[str(date_i.month)+'-'+str(date_i.year)] = num_opened
            date_i = date_i+relativedelta(months=1)
        f, (axis) = plt.subplots(nrows=1,ncols=2,sharey=True,figsize = (20,10))
        pre_fg_avg= sum(list(pre_fg_inspection.values()))/len(list(pre_fg_inspection.values()))
        pre_fg_avg_line = [pre_fg_avg]* len(list(pre_fg_inspection.values()))
        
        axis[0].plot(list(pre_fg_inspection.keys()),list(pre_fg_inspection.values()))
        axis[0].plot(list(pre_fg_inspection.keys()),pre_fg_avg_line,label = "Average")
        axis[0].set_title("FY20")
        axis[0].tick_params(labelrotation = 30)
        ylim = max(max(list(pre_fg_inspection.values())),max(list(post_fg_inspection.values())))
        axis[0].set_ylim(0,ylim)
        post_fg_avg = sum(list(post_fg_inspection.values()))/len(list(post_fg_inspection.values()))
        post_fg_avg_line = [post_fg_avg]* len(list(post_fg_inspection.values()))
        axis[1].plot(list(post_fg_inspection.keys()),list(post_fg_inspection.values()))
        axis[1].plot(list(post_fg_inspection.keys()),post_fg_avg_line)
        axis[1].set_title("FY21")
        axis[1].tick_params(labelrotation = 30)
        f.suptitle(f"Number of {ff} NCs Opened by Month",fontsize = 'x-large')
        #f.tight_layout() 
        f.legend()
        f.savefig(save_to_path+'\\'+"Number of NCs Opened by Month")
        plt.show()

    def check_for_lot_number(self, lot):
        self.r_data = pd.read_excel(open(self.path,'rb'))
        self.data = self.r_data.iloc[4:]
        self.data = self.data.rename(columns = self.data.iloc[0,:])
        self.data = self.data.drop(4)
        date_fields = ['Created Date','Report Date','Occurrence Date','Closure Due Date','Closed Date']
        self.date_converter(self.data,date_fields)
        self.data['Product'] = self.tmpecc_remover(self.data['Product'])
        self.data = self.data[self.data["NC Number"] != self.data.iloc[-1,0]]
        return self.data['NC Number'][self.data['Lot Number'] == lot].unique()

        





if __name__ == '__main__':
    protocol()
